"""
Gera o arquivo Excel Bliss_Living_Unidades.xlsx com todas as unidades do
empreendimento Bliss Living.

Uso:
    python gerar_planilha_bliss.py
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = "Bliss_Living_Unidades.xlsx"

# ---------------------------------------------------------------------------
# Cabeçalho
# ---------------------------------------------------------------------------
HEADERS = [
    "Unidade",
    "Localização",
    "Tipologia",
    "Tipo",
    "Vínculo Matrícula",
    "Área Privativa",
    "Área Privativa Acessória",
    "Área Privativa Total",
    "Área de Uso Comum",
    "Área Real Total",
    "Fração Ideal (%)",
    "Matrícula",
]

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
EVEN_FILL    = PatternFill("solid", fgColor="EBF0FA")
ODD_FILL     = PatternFill("solid", fgColor="FFFFFF")

FMT_AREA  = "0.000"
FMT_FRAC  = '0.000000"%"'   # Excel interpreta a % literal; usamos texto com %

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_br(value_str):
    """Converte '2,350' → 2.35 (float)."""
    return float(str(value_str).replace(",", "."))


def hb_row(numero, localizacao, priv, ac, total_priv, comum, total, fracao, matricula_txt):
    return {
        "unidade":      f"Hobby Box {numero:02d}" if isinstance(numero, int) else f"Hobby Box {numero}",
        "localizacao":  localizacao,
        "tipologia":    "",
        "tipo":         "Hobby Box",
        "vinculo":      "",
        "priv":         _parse_br(priv),
        "ac":           _parse_br(ac),
        "total_priv":   _parse_br(total_priv),
        "comum":        _parse_br(comum),
        "total":        _parse_br(total),
        "fracao":       _parse_br(fracao),
        "matricula":    matricula_txt,
    }


def g_row(numero, localizacao, tipologia, priv, ac, total_priv, comum, total, fracao, vinculo, matricula_txt):
    nome = f"Garagem {numero}"
    return {
        "unidade":      nome,
        "localizacao":  localizacao,
        "tipologia":    tipologia,
        "tipo":         "Garagem",
        "vinculo":      vinculo,
        "priv":         _parse_br(priv),
        "ac":           _parse_br(ac),
        "total_priv":   _parse_br(total_priv),
        "comum":        _parse_br(comum),
        "total":        _parse_br(total),
        "fracao":       _parse_br(fracao),
        "matricula":    matricula_txt,
    }


def apto_row(numero_str, bloco, localizacao, tipologia, priv, ac, total_priv, comum, total, fracao, matricula_txt):
    sufixo = "-SUN" if bloco == "SUN" else "-SHN"
    return {
        "unidade":      f"Apto {numero_str}{sufixo}",
        "localizacao":  localizacao,
        "tipologia":    tipologia,
        "tipo":         "Apartamento",
        "vinculo":      "",
        "priv":         _parse_br(priv),
        "ac":           _parse_br(ac),
        "total_priv":   _parse_br(total_priv),
        "comum":        _parse_br(comum),
        "total":        _parse_br(total),
        "fracao":       _parse_br(fracao),
        "matricula":    matricula_txt,
    }


def loja_row(priv, ac, total_priv, comum, total, fracao, matricula_txt):
    return {
        "unidade":      "Loja Comercial",
        "localizacao":  "Térreo",
        "tipologia":    "Com Mezanino/Sobreloja",
        "tipo":         "Loja",
        "vinculo":      "",
        "priv":         _parse_br(priv),
        "ac":           _parse_br(ac),
        "total_priv":   _parse_br(total_priv),
        "comum":        _parse_br(comum),
        "total":        _parse_br(total),
        "fracao":       _parse_br(fracao),
        "matricula":    matricula_txt,
    }


# ---------------------------------------------------------------------------
# Dados
# ---------------------------------------------------------------------------

def build_rows():
    rows = []
    LOC_SUB = "Subsolo"
    LOC_TER = "Térreo"
    LOC_G1  = "G1 - Shine"
    LOC_PIL = "Pilotis - Shine"

    # ========================================================================
    # SUBSOLO — Hobby Boxes autônomos
    # ========================================================================
    # HB 01, 21, 22, 23
    mat_hb_01 = (
        "os hobby box n°s 01, 21, 22 e 23 terão cada um a área privativa de 2,350m², "
        "área de uso comum de 4,179m², área total de 6,529m² e 0,035740% de fração ideal no terreno"
    )
    for n in [1, 21, 22, 23]:
        rows.append(hb_row(n, LOC_SUB, "2,350", "0", "2,350", "4,179", "6,529", "0,035740", mat_hb_01))

    # HB 02, 03, 04, 05, 07, 08, 11, 12
    mat_hb_02 = (
        "os hobby box n°s 02, 03, 04, 05, 07, 08, 11 e 12 terão cada um a área privativa de 2,190m², "
        "área de uso comum de 3,894m², área total de 6,084m² e 0,033313% de fração ideal no terreno"
    )
    for n in [2, 3, 4, 5, 7, 8, 11, 12]:
        rows.append(hb_row(n, LOC_SUB, "2,190", "0", "2,190", "3,894", "6,084", "0,033313", mat_hb_02))

    # HB 06, 25, 26
    mat_hb_06 = (
        "os hobby box n°s 06, 25 e 26 terão cada um a área privativa de 2,030m², "
        "área de uso comum de 3,610m², área total de 5,640m² e 0,030873% de fração ideal no terreno"
    )
    for n in [6, 25, 26]:
        rows.append(hb_row(n, LOC_SUB, "2,030", "0", "2,030", "3,610", "5,640", "0,030873", mat_hb_06))

    # HB 09, 10, 17, 18, 19, 20
    mat_hb_09 = (
        "os hobby box n°s 09, 10, 17, 18, 19 e 20 terão cada um a área privativa de 3,000m², "
        "área de uso comum de 5,335m², área total de 8,335m² e 0,045629% de fração ideal no terreno"
    )
    for n in [9, 10, 17, 18, 19, 20]:
        rows.append(hb_row(n, LOC_SUB, "3,000", "0", "3,000", "5,335", "8,335", "0,045629", mat_hb_09))

    # HB 13, 14, 15
    mat_hb_13 = (
        "os hobby box n°s 13, 14 e 15 terão cada um a área privativa de 2,700m², "
        "área de uso comum de 4,803m², área total de 7,503m² e 0,041065% de fração ideal no terreno"
    )
    for n in [13, 14, 15]:
        rows.append(hb_row(n, LOC_SUB, "2,700", "0", "2,700", "4,803", "7,503", "0,041065", mat_hb_13))

    # HB 16
    mat_hb_16 = (
        "o hobby box n° 16 terá a área privativa de 2,500m², "
        "área de uso comum de 4,448m², área total de 6,948m² e 0,038028% de fração ideal no terreno"
    )
    rows.append(hb_row(16, LOC_SUB, "2,500", "0", "2,500", "4,448", "6,948", "0,038028", mat_hb_16))

    # HB 32
    mat_hb_32 = (
        "o hobby box n° 32 terá a área privativa de 3,440m², "
        "área de uso comum de 6,118m², área total de 9,558m² e 0,052316% de fração ideal no terreno"
    )
    rows.append(hb_row(32, LOC_SUB, "3,440", "0", "3,440", "6,118", "9,558", "0,052316", mat_hb_32))

    # ========================================================================
    # SUBSOLO — Garagens
    # ========================================================================
    # G 09,10,11,12,15,16,17,18,19,20,24,25,27,28,29,30,33 (especiais, sem HB)
    mat_g_esp = (
        "as vagas de garagem n°s 09, 10, 11, 12, 15, 16, 17, 18, 19, 20, 24, 25, 27, 28, 29, 30 e 33 "
        "(especiais, sem hobby box) terão cada uma a área privativa de 21,600m², "
        "área de uso comum de 38,415m², área total de 60,015m² e 0,328521% de fração ideal no terreno"
    )
    for n in [9, 10, 11, 12, 15, 16, 17, 18, 19, 20, 24, 25, 27, 28, 29, 30, 33]:
        rows.append(g_row(f"{n:02d}", LOC_SUB, "Especial",
                          "21,600", "0", "21,600", "38,415", "60,015", "0,328521", "", mat_g_esp))

    # G 13, 14, 23, 26 (regular, sem HB)
    mat_g_reg = (
        "as vagas de garagem n°s 13, 14, 23 e 26 terão cada uma a área privativa de 12,000m², "
        "área de uso comum de 21,342m², área total de 33,342m² e 0,182518% de fração ideal no terreno"
    )
    for n in [13, 14, 23, 26]:
        rows.append(g_row(f"{n:02d}", LOC_SUB, "",
                          "12,000", "0", "12,000", "21,342", "33,342", "0,182518", "", mat_g_reg))

    # G 21 (com HB 27)
    rows.append(g_row("21", LOC_SUB, "Especial",
                      "21,600", "2,030", "23,630", "42,025", "65,655", "0,359394", "HB 27",
                      "a vaga de garagem n° 21 (com hobby box n° 27) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 2,030m², área privativa total de 23,630m², "
                      "área de uso comum de 42,025m², área total de 65,655m² e 0,359394% de fração ideal no terreno"))

    # G 22 (com HB 24)
    rows.append(g_row("22", LOC_SUB, "",
                      "12,000", "2,030", "14,030", "24,953", "38,983", "0,213391", "HB 24",
                      "a vaga de garagem n° 22 (com hobby box n° 24) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,030m², área privativa total de 14,030m², "
                      "área de uso comum de 24,953m², área total de 38,983m² e 0,213391% de fração ideal no terreno"))

    # G 31 (com HB 33)
    rows.append(g_row("31", LOC_SUB, "Especial",
                      "21,600", "3,440", "25,040", "44,532", "69,572", "0,380836", "HB 33",
                      "a vaga de garagem n° 31 (com hobby box n° 33) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 3,440m², área privativa total de 25,040m², "
                      "área de uso comum de 44,532m², área total de 69,572m² e 0,380836% de fração ideal no terreno"))

    # G 32 (com HB 34)
    rows.append(g_row("32", LOC_SUB, "Especial",
                      "21,600", "3,440", "25,040", "44,532", "69,572", "0,380836", "HB 34",
                      "a vaga de garagem n° 32 (com hobby box n° 34) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 3,440m², área privativa total de 25,040m², "
                      "área de uso comum de 44,532m², área total de 69,572m² e 0,380836% de fração ideal no terreno"))

    # G 34 (com HB 31)
    rows.append(g_row("34", LOC_SUB, "Especial",
                      "21,600", "4,480", "26,080", "46,382", "72,462", "0,396660", "HB 31",
                      "a vaga de garagem n° 34 (com hobby box n° 31) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 4,480m², área privativa total de 26,080m², "
                      "área de uso comum de 46,382m², área total de 72,462m² e 0,396660% de fração ideal no terreno"))

    # G 35 (com HB 30)
    rows.append(g_row("35", LOC_SUB, "",
                      "12,000", "4,240", "16,240", "28,882", "45,122", "0,247008", "HB 30",
                      "a vaga de garagem n° 35 (com hobby box n° 30) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 4,240m², área privativa total de 16,240m², "
                      "área de uso comum de 28,882m², área total de 45,122m² e 0,247008% de fração ideal no terreno"))

    # G 36 (com HB 29)
    rows.append(g_row("36", LOC_SUB, "",
                      "12,000", "3,200", "15,200", "27,033", "42,233", "0,231185", "HB 29",
                      "a vaga de garagem n° 36 (com hobby box n° 29) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 3,200m², área privativa total de 15,200m², "
                      "área de uso comum de 27,033m², área total de 42,233m² e 0,231185% de fração ideal no terreno"))

    # G 37 (com HB 28)
    rows.append(g_row("37", LOC_SUB, "",
                      "12,000", "3,440", "15,440", "27,460", "42,900", "0,234833", "HB 28",
                      "a vaga de garagem n° 37 (com hobby box n° 28) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 3,440m², área privativa total de 15,440m², "
                      "área de uso comum de 27,460m², área total de 42,900m² e 0,234833% de fração ideal no terreno"))

    # ========================================================================
    # TÉRREO — Hobby Boxes autônomos
    # ========================================================================
    # HB 35, 40, 44
    mat_hb_35 = (
        "os hobby box n°s 35, 40 e 44 terão cada um a área privativa de 2,190m², "
        "área de uso comum de 3,894m², área total de 6,084m² e 0,033313% de fração ideal no terreno"
    )
    for n in [35, 40, 44]:
        rows.append(hb_row(n, LOC_TER, "2,190", "0", "2,190", "3,894", "6,084", "0,033313", mat_hb_35))

    # HB 36, 42, 43
    mat_hb_36 = (
        "os hobby box n°s 36, 42 e 43 terão cada um a área privativa de 2,030m², "
        "área de uso comum de 3,610m², área total de 5,640m² e 0,030873% de fração ideal no terreno"
    )
    for n in [36, 42, 43]:
        rows.append(hb_row(n, LOC_TER, "2,030", "0", "2,030", "3,610", "5,640", "0,030873", mat_hb_36))

    # HB 39, 41
    mat_hb_39 = (
        "os hobby box n°s 39 e 41 terão cada um a área privativa de 2,350m², "
        "área de uso comum de 4,179m², área total de 6,529m² e 0,035740% de fração ideal no terreno"
    )
    for n in [39, 41]:
        rows.append(hb_row(n, LOC_TER, "2,350", "0", "2,350", "4,179", "6,529", "0,035740", mat_hb_39))

    # ========================================================================
    # TÉRREO — Garagens
    # ========================================================================
    # G 38, 41, 42, 43, 44, 45 (especiais)
    mat_g_ter_esp = (
        "as vagas de garagem n°s 38, 41, 42, 43, 44 e 45 (especiais) terão cada uma a área privativa de 21,600m², "
        "área de uso comum de 38,415m², área total de 60,015m² e 0,328521% de fração ideal no terreno"
    )
    for n in [38, 41, 42, 43, 44, 45]:
        rows.append(g_row(f"{n:02d}", LOC_TER, "Especial",
                          "21,600", "0", "21,600", "38,415", "60,015", "0,328521", "", mat_g_ter_esp))

    # G 40 PNE
    rows.append(g_row("40 PNE", LOC_TER, "PNE",
                      "21,600", "0", "21,600", "38,415", "60,015", "0,328532", "",
                      "a vaga de garagem n° 40 PNE terá a área privativa de 21,600m², "
                      "área de uso comum de 38,415m², área total de 60,015m² e 0,328532% de fração ideal no terreno"))

    # G 39
    rows.append(g_row("39", LOC_TER, "",
                      "12,000", "0", "12,000", "21,342", "33,342", "0,182518", "",
                      "a vaga de garagem n° 39 terá a área privativa de 12,000m², "
                      "área de uso comum de 21,342m², área total de 33,342m² e 0,182518% de fração ideal no terreno"))

    # G 46 idoso (com HB 37)
    rows.append(g_row("46 Idoso", LOC_TER, "Idoso",
                      "12,000", "2,500", "14,500", "25,790", "40,290", "0,220546", "HB 37",
                      "a vaga de garagem n° 46 idoso (com hobby box n° 37) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,500m², área privativa total de 14,500m², "
                      "área de uso comum de 25,790m², área total de 40,290m² e 0,220546% de fração ideal no terreno"))

    # G 47 idoso (com HB 38)
    rows.append(g_row("47 Idoso", LOC_TER, "Idoso",
                      "12,000", "3,440", "15,440", "27,460", "42,900", "0,234833", "HB 38",
                      "a vaga de garagem n° 47 idoso (com hobby box n° 38) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 3,440m², área privativa total de 15,440m², "
                      "área de uso comum de 27,460m², área total de 42,900m² e 0,234833% de fração ideal no terreno"))

    # G 48, 49, 50 idoso
    mat_g_idoso = (
        "as vagas de garagem n°s 48, 49 e 50 idoso terão cada uma a área privativa de 12,000m², "
        "área de uso comum de 21,342m², área total de 33,342m² e 0,182518% de fração ideal no terreno"
    )
    for n, label in [(48, "48 Idoso"), (49, "49 Idoso"), (50, "50 Idoso")]:
        rows.append(g_row(label, LOC_TER, "Idoso",
                          "12,000", "0", "12,000", "21,342", "33,342", "0,182518", "", mat_g_idoso))

    # G 51 PNE
    rows.append(g_row("51 PNE", LOC_TER, "PNE",
                      "22,500", "0", "22,500", "40,016", "62,516", "0,342220", "",
                      "a vaga de garagem n° 51 PNE terá a área privativa de 22,500m², "
                      "área de uso comum de 40,016m², área total de 62,516m² e 0,342220% de fração ideal no terreno"))

    # G 52/52A descoberta
    rows.append(g_row("52/52A", LOC_TER, "Dupla",
                      "12,000", "12,000", "24,000", "40,980", "64,980", "0,304192", "",
                      "a vaga de garagem n° 52/52A descoberta terá a área privativa de 12,000m², "
                      "área privativa acessória de 12,000m², área privativa total de 24,000m², "
                      "área de uso comum de 40,980m², área total de 64,980m² e 0,304192% de fração ideal no terreno"))

    # G 53/53A descoberta
    rows.append(g_row("53/53A", LOC_TER, "Dupla",
                      "12,000", "12,000", "24,000", "40,980", "64,980", "0,304192", "",
                      "a vaga de garagem n° 53/53A descoberta terá a área privativa de 12,000m², "
                      "área privativa acessória de 12,000m², área privativa total de 24,000m², "
                      "área de uso comum de 40,980m², área total de 64,980m² e 0,304192% de fração ideal no terreno"))

    # ========================================================================
    # TÉRREO — Loja
    # ========================================================================
    rows.append(loja_row(
        "499,290", "96,000", "595,290", "322,950", "918,240", "6,897441",
        "a loja comercial (com vagas de garagem n°s 01 a 08 no subsolo) terá a área privativa de 499,290m², "
        "área privativa acessória de 96,000m² (vagas G01 a G08), área privativa total de 595,290m², "
        "área de uso comum de 322,950m², área total de 918,240m² e 6,897441% de fração ideal no terreno"
    ))

    # ========================================================================
    # G1 — BLOCO 02 SHINE — Hobby Boxes autônomos
    # ========================================================================
    # HB 49, 50, 52, 53, 54, 61
    mat_hb_49 = (
        "os hobby box n°s 49, 50, 52, 53, 54 e 61 terão cada um a área privativa de 2,190m², "
        "área de uso comum de 3,894m², área total de 6,084m² e 0,033313% de fração ideal no terreno"
    )
    for n in [49, 50, 52, 53, 54, 61]:
        rows.append(hb_row(n, LOC_G1, "2,190", "0", "2,190", "3,894", "6,084", "0,033313", mat_hb_49))

    # HB 55
    rows.append(hb_row(55, LOC_G1, "2,030", "0", "2,030", "3,610", "5,640", "0,030873",
                       "o hobby box n° 55 terá a área privativa de 2,030m², "
                       "área de uso comum de 3,610m², área total de 5,640m² e 0,030873% de fração ideal no terreno"))

    # HB 60, 62
    mat_hb_60 = (
        "os hobby box n°s 60 e 62 terão cada um a área privativa de 2,350m², "
        "área de uso comum de 4,179m², área total de 6,529m² e 0,035740% de fração ideal no terreno"
    )
    for n in [60, 62]:
        rows.append(hb_row(n, LOC_G1, "2,350", "0", "2,350", "4,179", "6,529", "0,035740", mat_hb_60))

    # HB 63
    rows.append(hb_row(63, LOC_G1, "3,200", "0", "3,200", "5,691", "8,891", "0,048667",
                       "o hobby box n° 63 terá a área privativa de 3,200m², "
                       "área de uso comum de 5,691m², área total de 8,891m² e 0,048667% de fração ideal no terreno"))

    # ========================================================================
    # G1 — BLOCO 02 SHINE — Garagens
    # ========================================================================
    # G 54, 57, 58, 61, 62
    mat_g1_reg = (
        "as vagas de garagem n°s 54, 57, 58, 61 e 62 terão cada uma a área privativa de 12,000m², "
        "área de uso comum de 21,342m², área total de 33,342m² e 0,182518% de fração ideal no terreno"
    )
    for n in [54, 57, 58, 61, 62]:
        rows.append(g_row(f"{n:02d}", LOC_G1, "",
                          "12,000", "0", "12,000", "21,342", "33,342", "0,182518", "", mat_g1_reg))

    # G 68, 71, 72, 73, 74, 75 (especiais)
    mat_g1_esp = (
        "as vagas de garagem n°s 68, 71, 72, 73, 74 e 75 (especiais) terão cada uma a área privativa de 21,600m², "
        "área de uso comum de 38,415m², área total de 60,015m² e 0,328521% de fração ideal no terreno"
    )
    for n in [68, 71, 72, 73, 74, 75]:
        rows.append(g_row(f"{n:02d}", LOC_G1, "Especial",
                          "21,600", "0", "21,600", "38,415", "60,015", "0,328521", "", mat_g1_esp))

    # G 55 (com HB 45)
    rows.append(g_row("55", LOC_G1, "",
                      "12,000", "2,700", "14,700", "26,145", "40,845", "0,223584", "HB 45",
                      "a vaga de garagem n° 55 (com hobby box n° 45) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,700m², área privativa total de 14,700m², "
                      "área de uso comum de 26,145m², área total de 40,845m² e 0,223584% de fração ideal no terreno"))

    # G 56 (com HB 46)
    rows.append(g_row("56", LOC_G1, "",
                      "12,000", "2,700", "14,700", "26,145", "40,845", "0,223584", "HB 46",
                      "a vaga de garagem n° 56 (com hobby box n° 46) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,700m², área privativa total de 14,700m², "
                      "área de uso comum de 26,145m², área total de 40,845m² e 0,223584% de fração ideal no terreno"))

    # G 59 (com HB 47)
    rows.append(g_row("59", LOC_G1, "",
                      "12,000", "2,190", "14,190", "25,237", "39,427", "0,215830", "HB 47",
                      "a vaga de garagem n° 59 (com hobby box n° 47) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,190m², área privativa total de 14,190m², "
                      "área de uso comum de 25,237m², área total de 39,427m² e 0,215830% de fração ideal no terreno"))

    # G 60 (com HB 48)
    rows.append(g_row("60", LOC_G1, "",
                      "12,000", "2,190", "14,190", "25,237", "39,427", "0,215830", "HB 48",
                      "a vaga de garagem n° 60 (com hobby box n° 48) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,190m², área privativa total de 14,190m², "
                      "área de uso comum de 25,237m², área total de 39,427m² e 0,215830% de fração ideal no terreno"))

    # G 63 (com HB 58)
    rows.append(g_row("63", LOC_G1, "",
                      "12,000", "2,350", "14,350", "25,522", "39,872", "0,218259", "HB 58",
                      "a vaga de garagem n° 63 (com hobby box n° 58) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,350m², área privativa total de 14,350m², "
                      "área de uso comum de 25,522m², área total de 39,872m² e 0,218259% de fração ideal no terreno"))

    # G 64 (com HB 51)
    rows.append(g_row("64", LOC_G1, "",
                      "12,000", "2,190", "14,190", "25,237", "39,427", "0,215830", "HB 51",
                      "a vaga de garagem n° 64 (com hobby box n° 51) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,190m², área privativa total de 14,190m², "
                      "área de uso comum de 25,237m², área total de 39,427m² e 0,215830% de fração ideal no terreno"))

    # G 65 (com HB 56)
    rows.append(g_row("65", LOC_G1, "",
                      "12,000", "2,500", "14,500", "25,790", "40,290", "0,220546", "HB 56",
                      "a vaga de garagem n° 65 (com hobby box n° 56) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,500m², área privativa total de 14,500m², "
                      "área de uso comum de 25,790m², área total de 40,290m² e 0,220546% de fração ideal no terreno"))

    # G 66 (com HB 57)
    rows.append(g_row("66", LOC_G1, "",
                      "12,000", "3,440", "15,440", "27,460", "42,900", "0,234833", "HB 57",
                      "a vaga de garagem n° 66 (com hobby box n° 57) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 3,440m², área privativa total de 15,440m², "
                      "área de uso comum de 27,460m², área total de 42,900m² e 0,234833% de fração ideal no terreno"))

    # G 67/67A (com HB 59)
    rows.append(g_row("67/67A", LOC_G1, "Dupla",
                      "12,000", "14,350", "26,350", "46,865", "73,215", "0,400776", "HB 59",
                      "a vaga de garagem n° 67/67A (com hobby box n° 59) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 14,350m² (garagem 12,000m² + HB 2,350m²), "
                      "área privativa total de 26,350m², área de uso comum de 46,865m², "
                      "área total de 73,215m² e 0,400776% de fração ideal no terreno"))

    # G 69 especial (com HB 64)
    rows.append(g_row("69", LOC_G1, "Especial",
                      "21,600", "3,000", "24,600", "43,750", "68,350", "0,374150", "HB 64",
                      "a vaga de garagem n° 69 especial (com hobby box n° 64) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 3,000m², área privativa total de 24,600m², "
                      "área de uso comum de 43,750m², área total de 68,350m² e 0,374150% de fração ideal no terreno"))

    # G 70 especial (com HB 65)
    rows.append(g_row("70", LOC_G1, "Especial",
                      "21,600", "3,000", "24,600", "43,750", "68,350", "0,374150", "HB 65",
                      "a vaga de garagem n° 70 especial (com hobby box n° 65) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 3,000m², área privativa total de 24,600m², "
                      "área de uso comum de 43,750m², área total de 68,350m² e 0,374150% de fração ideal no terreno"))

    # ========================================================================
    # PILOTIS — BLOCO 02 SHINE — Hobby Boxes autônomos
    # ========================================================================
    # HB 66, 67, 68, 69, 70, 77
    mat_hb_66 = (
        "os hobby box n°s 66, 67, 68, 69, 70 e 77 terão cada um a área privativa de 2,190m², "
        "área de uso comum de 3,894m², área total de 6,084m² e 0,033313% de fração ideal no terreno"
    )
    for n in [66, 67, 68, 69, 70, 77]:
        rows.append(hb_row(n, LOC_PIL, "2,190", "0", "2,190", "3,894", "6,084", "0,033313", mat_hb_66))

    # HB 71
    rows.append(hb_row(71, LOC_PIL, "2,030", "0", "2,030", "3,610", "5,640", "0,030873",
                       "o hobby box n° 71 terá a área privativa de 2,030m², "
                       "área de uso comum de 3,610m², área total de 5,640m² e 0,030873% de fração ideal no terreno"))

    # HB 76, 78
    mat_hb_76 = (
        "os hobby box n°s 76 e 78 terão cada um a área privativa de 2,350m², "
        "área de uso comum de 4,179m², área total de 6,529m² e 0,035740% de fração ideal no terreno"
    )
    for n in [76, 78]:
        rows.append(hb_row(n, LOC_PIL, "2,350", "0", "2,350", "4,179", "6,529", "0,035740", mat_hb_76))

    # HB 79
    rows.append(hb_row(79, LOC_PIL, "3,200", "0", "3,200", "5,691", "8,891", "0,048667",
                       "o hobby box n° 79 terá a área privativa de 3,200m², "
                       "área de uso comum de 5,691m², área total de 8,891m² e 0,048667% de fração ideal no terreno"))

    # ========================================================================
    # PILOTIS — BLOCO 02 SHINE — Garagens
    # ========================================================================
    # G 81
    rows.append(g_row("81", LOC_PIL, "",
                      "12,000", "0", "12,000", "21,342", "33,342", "0,182518", "",
                      "a vaga de garagem n° 81 terá a área privativa de 12,000m², "
                      "área de uso comum de 21,342m², área total de 33,342m² e 0,182518% de fração ideal no terreno"))

    # G 77, 78, 83 (descobertas)
    mat_g_pil_desc = (
        "as vagas de garagem n°s 77, 78 e 83 (descobertas) terão cada uma a área privativa de 12,000m², "
        "área de uso comum de 19,638m², área total de 31,638m² e 0,121675% de fração ideal no terreno"
    )
    for n in [77, 78, 83]:
        rows.append(g_row(f"{n:02d}", LOC_PIL, "Descoberta",
                          "12,000", "0", "12,000", "19,638", "31,638", "0,121675", "", mat_g_pil_desc))

    # G 87, 93, 94 (especiais descobertas)
    mat_g_pil_esp_desc = (
        "as vagas de garagem n°s 87, 93 e 94 (especiais descobertas) terão cada uma a área privativa de 21,600m², "
        "área de uso comum de 35,349m², área total de 56,949m² e 0,219021% de fração ideal no terreno"
    )
    for n in [87, 93, 94]:
        rows.append(g_row(f"{n:02d}", LOC_PIL, "Especial Descoberta",
                          "21,600", "0", "21,600", "35,349", "56,949", "0,219021", "", mat_g_pil_esp_desc))

    # G 90 (especial parc. descob.)
    rows.append(g_row("90", LOC_PIL, "Especial Parcialmente Descoberta",
                      "21,600", "0", "21,600", "37,800", "59,400", "0,306577", "",
                      "a vaga de garagem n° 90 (especial parcialmente descoberta) terá a área privativa de 21,600m², "
                      "área de uso comum de 37,800m², área total de 59,400m² e 0,306577% de fração ideal no terreno"))

    # G 91 (especial parc. descob.)
    rows.append(g_row("91", LOC_PIL, "Especial Parcialmente Descoberta",
                      "21,600", "0", "21,600", "37,290", "58,890", "0,288325", "",
                      "a vaga de garagem n° 91 (especial parcialmente descoberta) terá a área privativa de 21,600m², "
                      "área de uso comum de 37,290m², área total de 58,890m² e 0,288325% de fração ideal no terreno"))

    # G 92 (especial parc. descob.)
    rows.append(g_row("92", LOC_PIL, "Especial Parcialmente Descoberta",
                      "21,600", "0", "21,600", "36,705", "58,305", "0,267438", "",
                      "a vaga de garagem n° 92 (especial parcialmente descoberta) terá a área privativa de 21,600m², "
                      "área de uso comum de 36,705m², área total de 58,305m² e 0,267438% de fração ideal no terreno"))

    # G 76/76A
    rows.append(g_row("76/76A", LOC_PIL, "Dupla",
                      "12,000", "12,000", "24,000", "40,980", "64,980", "0,304192", "",
                      "a vaga de garagem n° 76/76A terá a área privativa de 12,000m², "
                      "área privativa acessória de 12,000m², área privativa total de 24,000m², "
                      "área de uso comum de 40,980m², área total de 64,980m² e 0,304192% de fração ideal no terreno"))

    # G 79/79A
    rows.append(g_row("79/79A", LOC_PIL, "Dupla",
                      "12,000", "12,000", "24,000", "42,685", "66,685", "0,365035", "",
                      "a vaga de garagem n° 79/79A terá a área privativa de 12,000m², "
                      "área privativa acessória de 12,000m², área privativa total de 24,000m², "
                      "área de uso comum de 42,685m², área total de 66,685m² e 0,365035% de fração ideal no terreno"))

    # G 82/82A
    rows.append(g_row("82/82A", LOC_PIL, "Dupla",
                      "12,000", "12,000", "24,000", "40,980", "64,980", "0,304192", "",
                      "a vaga de garagem n° 82/82A terá a área privativa de 12,000m², "
                      "área privativa acessória de 12,000m², área privativa total de 24,000m², "
                      "área de uso comum de 40,980m², área total de 64,980m² e 0,304192% de fração ideal no terreno"))

    # G 80/80A (com HB 74)
    rows.append(g_row("80/80A", LOC_PIL, "Dupla",
                      "12,000", "14,350", "26,350", "46,865", "73,215", "0,400776", "HB 74",
                      "a vaga de garagem n° 80/80A (com hobby box n° 74) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 14,350m² (garagem 12,000m² + HB 2,350m²), "
                      "área privativa total de 26,350m², área de uso comum de 46,865m², "
                      "área total de 73,215m² e 0,400776% de fração ideal no terreno"))

    # G 84 (com HB 72)
    rows.append(g_row("84", LOC_PIL, "",
                      "12,000", "2,500", "14,500", "25,790", "40,290", "0,220546", "HB 72",
                      "a vaga de garagem n° 84 (com hobby box n° 72) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 2,500m², área privativa total de 14,500m², "
                      "área de uso comum de 25,790m², área total de 40,290m² e 0,220546% de fração ideal no terreno"))

    # G 85 (com HB 73)
    rows.append(g_row("85", LOC_PIL, "",
                      "12,000", "3,440", "15,440", "27,460", "42,900", "0,234833", "HB 73",
                      "a vaga de garagem n° 85 (com hobby box n° 73) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 3,440m², área privativa total de 15,440m², "
                      "área de uso comum de 27,460m², área total de 42,900m² e 0,234833% de fração ideal no terreno"))

    # G 86/86A (com HB 75)
    rows.append(g_row("86/86A", LOC_PIL, "Dupla",
                      "12,000", "14,350", "26,350", "46,865", "73,215", "0,400776", "HB 75",
                      "a vaga de garagem n° 86/86A (com hobby box n° 75) terá a área privativa principal de 12,000m², "
                      "área privativa acessória de 14,350m² (garagem 12,000m² + HB 2,350m²), "
                      "área privativa total de 26,350m², área de uso comum de 46,865m², "
                      "área total de 73,215m² e 0,400776% de fração ideal no terreno"))

    # G 88 especial descob. (com HB 80)
    rows.append(g_row("88", LOC_PIL, "Especial Descoberta",
                      "21,600", "3,000", "24,600", "40,684", "65,284", "0,264650", "HB 80",
                      "a vaga de garagem n° 88 especial descoberta (com hobby box n° 80) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 3,000m², área privativa total de 24,600m², "
                      "área de uso comum de 40,684m², área total de 65,284m² e 0,264650% de fração ideal no terreno"))

    # G 89 especial parc.descob. (com HB 81)
    rows.append(g_row("89", LOC_PIL, "Especial Parcialmente Descoberta",
                      "21,600", "3,000", "24,600", "43,670", "68,270", "0,371297", "HB 81",
                      "a vaga de garagem n° 89 especial parcialmente descoberta (com hobby box n° 81) terá a área privativa principal de 21,600m², "
                      "área privativa acessória de 3,000m², área privativa total de 24,600m², "
                      "área de uso comum de 43,670m², área total de 68,270m² e 0,371297% de fração ideal no terreno"))

    # ========================================================================
    # APARTAMENTOS — BLOCO 01 SUN — pavimentos 1 a 7
    # ========================================================================
    # Dados por coluna de apartamento (índice final: 1-6)
    sun_data = {
        "01": ("79,140", "0", "79,140", "24,126", "103,266", "0,861840"),
        "02": ("76,580", "0", "76,580", "23,346",  "99,926",  "0,833961"),
        "03": ("71,730", "0", "71,730", "21,867",  "93,597",  "0,781144"),
        "04": ("73,000", "0", "73,000", "22,255",  "95,255",  "0,794975"),
        "05": ("73,690", "0", "73,690", "22,465",  "96,155",  "0,802489"),
        "06": ("75,160", "0", "75,160", "22,913",  "98,073",  "0,818497"),
    }
    # Textos de matrícula — referem todos os andares
    sun_mat = {
        "01": (
            "os apartamentos n°s 101, 201, 301, 401, 501, 601 e 701 terão cada um a área privativa de 79,140m², "
            "área de uso comum de 24,126m², área total de 103,266m² e 0,861840% de fração ideal no terreno"
        ),
        "02": (
            "os apartamentos n°s 102, 202, 302, 402, 502, 602 e 702 terão cada um a área privativa de 76,580m², "
            "área de uso comum de 23,346m², área total de 99,926m² e 0,833961% de fração ideal no terreno"
        ),
        "03": (
            "os apartamentos n°s 103, 203, 303, 403, 503, 603 e 703 terão cada um a área privativa de 71,730m², "
            "área de uso comum de 21,867m², área total de 93,597m² e 0,781144% de fração ideal no terreno"
        ),
        "04": (
            "os apartamentos n°s 104, 204, 304, 404, 504, 604 e 704 terão cada um a área privativa de 73,000m², "
            "área de uso comum de 22,255m², área total de 95,255m² e 0,794975% de fração ideal no terreno"
        ),
        "05": (
            "os apartamentos n°s 105, 205, 305, 405, 505, 605 e 705 terão cada um a área privativa de 73,690m², "
            "área de uso comum de 22,465m², área total de 96,155m² e 0,802489% de fração ideal no terreno"
        ),
        "06": (
            "os apartamentos n°s 106, 206, 306, 406, 506, 606 e 706 terão cada um a área privativa de 75,160m², "
            "área de uso comum de 22,913m², área total de 98,073m² e 0,818497% de fração ideal no terreno"
        ),
    }
    # Sufixo ordinal para localização
    _ord = {1: "1°", 2: "2°", 3: "3°", 4: "4°", 5: "5°", 6: "6°", 7: "7°"}
    for pav in range(1, 8):
        loc_sun = f"{_ord[pav]} Pav. - Sun"
        for col in ["01", "02", "03", "04", "05", "06"]:
            numero = f"{pav}{col}"
            priv, ac, tp, com, tot, fra = sun_data[col]
            rows.append(apto_row(
                numero, "SUN", loc_sun, col,
                priv, ac, tp, com, tot, fra,
                sun_mat[col]
            ))

    # ========================================================================
    # APARTAMENTOS — BLOCO 02 SHINE — 1° PAVIMENTO
    # ========================================================================
    shine_1_data = {
        "101": ("73,800", "0", "73,800", "22,499", "96,299",  "0,803687"),
        "102": ("72,510", "0", "72,510", "22,105", "94,615",  "0,789639"),
        "103": ("72,550", "0", "72,550", "22,118", "94,668",  "0,790074"),
        "104": ("75,420", "0", "75,420", "22,992", "98,412",  "0,821329"),
        "105": ("70,390", "0", "70,390", "21,459", "91,849",  "0,766552"),
    }
    shine_1_mat = {
        "101": (
            "o apartamento n° 101-SHN terá a área privativa de 73,800m², "
            "área de uso comum de 22,499m², área total de 96,299m² e 0,803687% de fração ideal no terreno"
        ),
        "102": (
            "o apartamento n° 102-SHN terá a área privativa de 72,510m², "
            "área de uso comum de 22,105m², área total de 94,615m² e 0,789639% de fração ideal no terreno"
        ),
        "103": (
            "o apartamento n° 103-SHN terá a área privativa de 72,550m², "
            "área de uso comum de 22,118m², área total de 94,668m² e 0,790074% de fração ideal no terreno"
        ),
        "104": (
            "o apartamento n° 104-SHN terá a área privativa de 75,420m², "
            "área de uso comum de 22,992m², área total de 98,412m² e 0,821329% de fração ideal no terreno"
        ),
        "105": (
            "o apartamento n° 105-SHN terá a área privativa de 70,390m², "
            "área de uso comum de 21,459m², área total de 91,849m² e 0,766552% de fração ideal no terreno"
        ),
    }
    for apto_num, (priv, ac, tp, com, tot, fra) in shine_1_data.items():
        col_tip = apto_num[-2:]  # "01", "02", ...
        rows.append(apto_row(
            apto_num, "SHN", "1° Pav. - Shine", col_tip,
            priv, ac, tp, com, tot, fra,
            shine_1_mat[apto_num]
        ))

    # ========================================================================
    # APARTAMENTOS — BLOCO 02 SHINE — Pavimentos 2 a 7
    # ========================================================================
    shine_pav_data = {
        "01": ("73,800", "0", "73,800", "22,499", "96,299",  "0,803687"),
        "02": ("72,510", "0", "72,510", "22,105", "94,615",  "0,789639"),
        "03": ("72,550", "0", "72,550", "22,118", "94,668",  "0,790074"),
        "04": ("75,420", "0", "75,420", "22,992", "98,412",  "0,821329"),
        "05": ("70,390", "0", "70,390", "21,459", "91,849",  "0,766552"),
    }
    shine_pav_mat = {
        "01": (
            "os apartamentos n°s 201, 301, 401, 501, 601 e 701 (Shine) terão cada um a área privativa de 73,800m², "
            "área de uso comum de 22,499m², área total de 96,299m² e 0,803687% de fração ideal no terreno"
        ),
        "02": (
            "os apartamentos n°s 202, 302, 402, 502, 602 e 702 (Shine) terão cada um a área privativa de 72,510m², "
            "área de uso comum de 22,105m², área total de 94,615m² e 0,789639% de fração ideal no terreno"
        ),
        "03": (
            "os apartamentos n°s 203, 303, 403, 503, 603 e 703 (Shine) terão cada um a área privativa de 72,550m², "
            "área de uso comum de 22,118m², área total de 94,668m² e 0,790074% de fração ideal no terreno"
        ),
        "04": (
            "os apartamentos n°s 204, 304, 404, 504, 604 e 704 (Shine) terão cada um a área privativa de 75,420m², "
            "área de uso comum de 22,992m², área total de 98,412m² e 0,821329% de fração ideal no terreno"
        ),
        "05": (
            "os apartamentos n°s 205, 305, 405, 505, 605 e 705 (Shine) terão cada um a área privativa de 70,390m², "
            "área de uso comum de 21,459m², área total de 91,849m² e 0,766552% de fração ideal no terreno"
        ),
    }
    for pav in range(2, 8):
        loc_shine = f"{_ord[pav]} Pav. - Shine"
        for col in ["01", "02", "03", "04", "05"]:
            numero = f"{pav}{col}"
            priv, ac, tp, com, tot, fra = shine_pav_data[col]
            rows.append(apto_row(
                numero, "SHN", loc_shine, col,
                priv, ac, tp, com, tot, fra,
                shine_pav_mat[col]
            ))

    # ========================================================================
    # PAVIMENTO ÁTICO
    # ========================================================================
    # 801-SUN (cobertura)
    rows.append(apto_row(
        "801", "SUN", "Pav. Ático - Sun", "01 (Cobertura)",
        "180,130", "0", "180,130", "36,881", "217,011", "1,317458",
        "o apartamento n° 801-SUN (cobertura) terá a área privativa de 180,130m², "
        "área de uso comum de 36,881m², área total de 217,011m² e 1,317458% de fração ideal no terreno"
    ))
    # 802-SUN (cobertura)
    rows.append(apto_row(
        "802", "SUN", "Pav. Ático - Sun", "02 (Cobertura)",
        "217,260", "0", "217,260", "43,166", "260,426", "1,541957",
        "o apartamento n° 802-SUN (cobertura) terá a área privativa de 217,260m², "
        "área de uso comum de 43,166m², área total de 260,426m² e 1,541957% de fração ideal no terreno"
    ))
    # 801-SHN (cobertura)
    rows.append(apto_row(
        "801", "SHN", "Pav. Ático - Shine", "01 (Cobertura)",
        "158,150", "0", "158,150", "31,477", "189,627", "1,124399",
        "o apartamento n° 801-SHN (cobertura) terá a área privativa de 158,150m², "
        "área de uso comum de 31,477m², área total de 189,627m² e 1,124399% de fração ideal no terreno"
    ))
    # 802-SHN (cobertura)
    rows.append(apto_row(
        "802", "SHN", "Pav. Ático - Shine", "02 (Cobertura)",
        "176,270", "0", "176,270", "33,176", "209,446", "1,185090",
        "o apartamento n° 802-SHN (cobertura) terá a área privativa de 176,270m², "
        "área de uso comum de 33,176m², área total de 209,446m² e 1,185090% de fração ideal no terreno"
    ))

    return rows


# ---------------------------------------------------------------------------
# Geração do Excel
# ---------------------------------------------------------------------------

def gerar_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Unidades"

    # --- Cabeçalho ---
    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Dados ---
    NUM_COLS = [6, 7, 8, 9, 10]   # 1-indexed: Área Privativa … Área Real Total
    FRAC_COL = 11                  # Fração Ideal
    MAT_COL  = 12                  # Matrícula

    for row_idx, r in enumerate(rows, start=2):
        data_row = [
            r["unidade"],
            r["localizacao"],
            r["tipologia"],
            r["tipo"],
            r["vinculo"],
            r["priv"],
            r["ac"],
            r["total_priv"],
            r["comum"],
            r["total"],
            r["fracao"] / 100,     # fração em % → armazenamos como decimal para o formato %
            r["matricula"],
        ]
        ws.append(data_row)

        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == MAT_COL))
            if col_idx in NUM_COLS:
                cell.number_format = "0.000"
            elif col_idx == FRAC_COL:
                cell.number_format = '0.000000%'
            elif col_idx == MAT_COL:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # --- Larguras de coluna ---
    col_widths = {
        1:  30,   # Unidade
        2:  22,   # Localização
        3:  35,   # Tipologia
        4:  16,   # Tipo
        5:  18,   # Vínculo Matrícula
        6:  18,   # Área Privativa
        7:  22,   # Área Privativa Acessória
        8:  22,   # Área Privativa Total
        9:  20,   # Área de Uso Comum
        10: 18,   # Área Real Total
        11: 18,   # Fração Ideal
        12: 60,   # Matrícula
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Altura do cabeçalho ---
    ws.row_dimensions[1].height = 40

    # --- Congelar primeira linha ---
    ws.freeze_panes = "A2"

    # --- AutoFilter ---
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Arquivo salvo: {output_path}")
    return len(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import os

    output_path = os.path.join(os.path.dirname(__file__), "Bliss_Living_Unidades.xlsx")
    rows = build_rows()
    total = gerar_excel(rows, output_path)
    print(f"Total de unidades geradas: {total}")

    # Contagem por tipo
    tipos = {}
    for r in rows:
        tipos[r["tipo"]] = tipos.get(r["tipo"], 0) + 1
    for tipo, qtd in sorted(tipos.items()):
        print(f"  {tipo}: {qtd}")
